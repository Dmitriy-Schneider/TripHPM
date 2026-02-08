"""
Сервис для генерации документов (Word и Excel)
"""
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime, time
from typing import Dict, List
import zipfile
from num2words import num2words
import shutil
import logging

# Настройка логирования
logger = logging.getLogger(__name__)


class DocumentGenerator:
    """Генератор документов командировки"""

    def __init__(self, templates_dir: Path, output_dir: Path):
        self.templates_dir = Path(templates_dir)
        self.output_dir = Path(output_dir)

        # Валидация наличия шаблонов при инициализации
        self._validate_templates()

    def _validate_templates(self):
        """Проверяет наличие всех необходимых шаблонов"""
        required_templates = [
            "prikaz_template.docx",
            "ao_template.xlsx",
            "sz_template.docx"
        ]

        missing = []
        for template in required_templates:
            if not (self.templates_dir / template).exists():
                missing.append(template)

        if missing:
            logger.error(f"Missing templates: {', '.join(missing)}")
            raise FileNotFoundError(f"Required templates not found: {', '.join(missing)}")

        logger.info("[OK] All templates validated successfully")

    def generate_all(self, trip_data: Dict) -> Dict[str, str]:
        """
        Генерирует все документы командировки
        Возвращает словарь {doc_type: file_path}
        """
        logger.info(f"[GENERATE] Starting document generation for trip to {trip_data.get('destination_city')}")

        try:
            # Создаем папку для командировки
            trip_folder = self._create_trip_folder(trip_data)
            logger.info(f"[GENERATE] Created trip folder: {trip_folder}")

            results = {}

            # 1. Приказ
            logger.info("[GENERATE] Generating Prikaz...")
            results['prikaz'] = self._generate_prikaz(trip_data, trip_folder)
            logger.info(f"[GENERATE] Prikaz completed: {results['prikaz']}")

            # 2. Авансовый отчет (включает расчет суточных)
            logger.info("[GENERATE] Generating Avansoviy Otchet...")
            results['ao'] = self._generate_ao(trip_data, trip_folder)
            logger.info(f"[GENERATE] AO completed: {results['ao']}")

            # 3. Служебная записка
            logger.info("[GENERATE] Generating Sluzhebnaya Zapiska...")
            results['sz'] = self._generate_sz(trip_data, trip_folder)
            logger.info(f"[GENERATE] SZ completed: {results['sz']}")

            # 4. Копируем чеки
            logger.info("[GENERATE] Copying receipts...")
            self._copy_receipts(trip_data, trip_folder)
            logger.info(f"[GENERATE] Copied {len(trip_data.get('receipts', []))} receipts")

            # 5. Создаем ZIP архив
            logger.info("[GENERATE] Creating ZIP archive...")
            results['zip'] = self._create_zip(trip_folder, trip_data)
            logger.info(f"[GENERATE] ZIP completed: {results['zip']}")

            logger.info("[GENERATE] All documents generated successfully!")
            return results

        except Exception as e:
            logger.error(f"[GENERATE] CRITICAL ERROR: {str(e)}", exc_info=True)
            raise

    def _create_trip_folder(self, trip_data: Dict) -> Path:
        """Создает папку для командировки с полной очисткой"""
        folder_name = f"{trip_data['date_from'].strftime('%Y-%m-%d')}_{trip_data['destination_city']}_{trip_data['fio'].split()[0]}"
        trip_folder = self.output_dir / folder_name

        # КРИТИЧНО: Удаляем старую папку полностью перед созданием новой
        if trip_folder.exists():
            logger.warning(f"[FOLDER] Removing old folder: {trip_folder}")
            shutil.rmtree(trip_folder)
            logger.info(f"[FOLDER] Old folder removed successfully")

        # Создаем чистые папки
        trip_folder.mkdir(parents=True, exist_ok=False)
        (trip_folder / "documents").mkdir(exist_ok=False)
        (trip_folder / "receipts").mkdir(exist_ok=False)

        logger.info(f"[FOLDER] Created clean folder structure: {trip_folder}")

        return trip_folder

    def _generate_prikaz(self, trip_data: Dict, output_folder: Path) -> str:
        """Генерирует Приказ на командировку (Word)"""
        template_path = self.templates_dir / "prikaz_template.docx"
        output_path = output_folder / "documents" / "Приказ.docx"

        try:
            doc = DocxTemplate(template_path)

            context = {
                'org_name': trip_data.get('org_name', 'ООО «ВЭМ»'),
                'fio': trip_data['fio'],
                'tab_no': trip_data.get('tab_no', ''),
                'position': trip_data.get('position', ''),
                'department': trip_data.get('department', ''),
                'destination_city': trip_data['destination_city'],
                'destination_org': trip_data['destination_org'],
                'date_from': trip_data['date_from'].strftime('%d.%m.%Y'),
                'date_to': trip_data['date_to'].strftime('%d.%m.%Y'),
                'days': trip_data['days'],
                'purpose': trip_data['purpose'],
                'order_date': trip_data['date_from'].strftime('%d.%m.%Y'),
            }

            doc.render(context)
            doc.save(output_path)

            logger.info(f"[PRIKAZ] Generated successfully")
            return str(output_path)

        except Exception as e:
            logger.error(f"[PRIKAZ] Error: {str(e)}", exc_info=True)
            raise

    def _generate_ao(self, trip_data: Dict, output_folder: Path) -> str:
        """Генерирует Авансовый отчет (Excel)"""
        template_path = self.templates_dir / "ao_template.xlsx"
        output_path = output_folder / "documents" / "Авансовый_отчет.xlsx"

        try:
            # Открываем шаблон с сохранением формул
            wb = load_workbook(template_path, keep_vba=True, data_only=False, keep_links=True)
            ws = wb.active

            logger.info(f"[AO] Template loaded, sheet: {ws.title}")

            # Маппинг категорий на русские названия
            category_names = {
                'taxi': 'Такси',
                'fuel': 'Бензин',
                'airplane': 'Самолет',
                'flight': 'Самолет',
                'train': 'Поезд',
                'hotel': 'Гостиница',
                'restaurant': 'Ресторан',
                'other': 'Представительские'
            }

            # Расходы по категориям
            expenses = trip_data.get('expenses_by_category', {})
            logger.info(f"[AO] Processing {len(expenses)} expense categories: {list(expenses.keys())}")

            # Начальная строка для заполнения расходов (из скриншота пользователя)
            # Строка 63 - первая строка таблицы расходов
            start_row = 63
            current_row = start_row

            # Заполняем расходы по категориям
            for category, amount in expenses.items():
                if amount and amount > 0:
                    logger.info(f"[AO] Writing row {current_row}: {category} = {amount:.2f}")

                    try:
                        # Колонка P (16) - наименование документа
                        cell_p = ws.cell(current_row, 16)
                        cell_p.value = category_names.get(category, category)
                        logger.info(f"[AO]   P{current_row} = '{cell_p.value}'")
                    except Exception as e:
                        logger.error(f"[AO]   Error writing P{current_row}: {e}")

                    try:
                        # Колонка Y (25) - сумма в рублях
                        cell_y = ws.cell(current_row, 25)
                        cell_y.value = amount
                        logger.info(f"[AO]   Y{current_row} = {cell_y.value}")
                    except Exception as e:
                        logger.error(f"[AO]   Error writing Y{current_row}: {e}")

                    current_row += 1

            # Суточные
            per_diem_to_pay = trip_data.get('per_diem_to_pay', 0)
            if per_diem_to_pay > 0:
                logger.info(f"[AO] Writing per diem row {current_row}: {per_diem_to_pay:.2f}")

                try:
                    cell_p = ws.cell(current_row, 16)
                    cell_p.value = 'Суточные'
                    logger.info(f"[AO]   P{current_row} = 'Суточные'")
                except Exception as e:
                    logger.error(f"[AO]   Error writing per diem P{current_row}: {e}")

                try:
                    cell_y = ws.cell(current_row, 25)
                    cell_y.value = per_diem_to_pay
                    logger.info(f"[AO]   Y{current_row} = {per_diem_to_pay}")
                except Exception as e:
                    logger.error(f"[AO]   Error writing per diem Y{current_row}: {e}")

                current_row += 1
            else:
                logger.warning("[AO] Per diem is 0, skipping")

            wb.save(output_path)
            logger.info(f"[AO] Saved successfully to {output_path}")

            return str(output_path)

        except Exception as e:
            logger.error(f"[AO] CRITICAL ERROR: {str(e)}", exc_info=True)
            raise

    def _generate_sz(self, trip_data: Dict, output_folder: Path) -> str:
        """Генерирует Служебную записку (Word)"""
        template_path = self.templates_dir / "sz_template.docx"
        output_path = output_folder / "documents" / "Служебная_записка.docx"

        try:
            doc = DocxTemplate(template_path)

            total_expenses = trip_data.get('total_expenses', 0)
            to_return = trip_data.get('to_return', 0)

            context = {
                'fio': trip_data['fio'],
                'tab_no': trip_data.get('tab_no', ''),
                'trip_dates': f"{trip_data['date_from'].strftime('%d.%m.%Y')} - {trip_data['date_to'].strftime('%d.%m.%Y')}",
                'destination_city': trip_data['destination_city'],
                'total_expenses': f"{total_expenses:.2f}",
                'advance_rub': f"{trip_data.get('advance_rub', 0):.2f}",
                'to_return': f"{abs(to_return):.2f}",
                'to_return_text': 'к возврату' if to_return > 0 else 'к доплате',
                'sum_in_words': num2words(abs(to_return), lang='ru', to='currency', currency='RUB')
            }

            doc.render(context)
            doc.save(output_path)

            logger.info(f"[SZ] Generated successfully")
            return str(output_path)

        except Exception as e:
            logger.error(f"[SZ] Error: {str(e)}", exc_info=True)
            raise

    def _copy_receipts(self, trip_data: Dict, output_folder: Path):
        """Копирует чеки в папку командировки"""
        receipts_folder = output_folder / "receipts"
        receipts = trip_data.get('receipts', [])

        logger.info(f"[RECEIPTS] Starting to copy {len(receipts)} receipts")

        # Папка уже создана в _create_trip_folder и гарантированно пустая

        if not receipts:
            logger.warning("[RECEIPTS] No receipts to copy")
            return

        copied_count = 0
        for receipt in receipts:
            if 'file_path' not in receipt:
                logger.warning(f"[RECEIPTS] Receipt missing file_path: {receipt}")
                continue

            source = Path(receipt['file_path'])

            if not source.exists():
                logger.warning(f"[RECEIPTS] File not found: {source}")
                continue

            dest = receipts_folder / f"{receipt.get('category', 'other')}_{source.name}"

            try:
                shutil.copy(source, dest)
                copied_count += 1
                logger.info(f"[RECEIPTS] Copied: {source.name} -> {dest.name}")
            except Exception as e:
                logger.error(f"[RECEIPTS] Failed to copy {source.name}: {e}")

        logger.info(f"[RECEIPTS] Successfully copied {copied_count}/{len(receipts)} files")

    def _create_zip(self, trip_folder: Path, trip_data: Dict) -> str:
        """Создает ZIP архив со всеми документами и чеками"""
        zip_path = trip_folder.parent / f"{trip_folder.name}.zip"

        try:
            # Удаляем старый ZIP если существует
            if zip_path.exists():
                zip_path.unlink()
                logger.info(f"[ZIP] Removed old archive: {zip_path}")

            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Добавляем все файлы из папки
                file_count = 0
                for file_path in trip_folder.rglob('*'):
                    if file_path.is_file():
                        arcname = file_path.relative_to(trip_folder.parent)
                        zipf.write(file_path, arcname)
                        file_count += 1

                logger.info(f"[ZIP] Added {file_count} files to archive")

            logger.info(f"[ZIP] Created successfully: {zip_path}")
            return str(zip_path)

        except Exception as e:
            logger.error(f"[ZIP] Error: {str(e)}", exc_info=True)
            raise


def calculate_per_diem_days(date_from, date_to, departure_time, arrival_time) -> float:
    """
    Рассчитывает количество дней для суточных с коэффициентами

    Правила:
    - Выезд до 12:00 → коэф 1.0
    - Выезд 12:00-18:00 → коэф 0.5
    - Выезд после 18:00 → коэф 0.4
    - Приезд после 18:00 → коэф 1.0
    - Приезд 12:00-18:00 → коэф 0.5
    - Приезд до 12:00 → коэф 0.4
    """
    # Коэффициент выезда
    if departure_time:
        if departure_time.hour < 12:
            departure_coef = 1.0
        elif 12 <= departure_time.hour < 18:
            departure_coef = 0.5
        else:
            departure_coef = 0.4
    else:
        departure_coef = 1.0

    # Коэффициент приезда
    if arrival_time:
        if arrival_time.hour >= 18:
            arrival_coef = 1.0
        elif 12 <= arrival_time.hour < 18:
            arrival_coef = 0.5
        else:
            arrival_coef = 0.4
    else:
        arrival_coef = 1.0

    # Полные дни между
    days_diff = (date_to - date_from).days
    full_days = max(0, days_diff - 1)

    total_days = departure_coef + full_days + arrival_coef

    logger.info(f"[PER_DIEM] Calculated: departure_coef={departure_coef}, full_days={full_days}, arrival_coef={arrival_coef}, total={total_days}")

    return total_days
